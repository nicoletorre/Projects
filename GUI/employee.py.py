from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox

from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook

excel_con = Workbook()
excel_con = load_workbook('finals.xlsx')
excel_activate = excel_con.active
column1= excel_activate['A']

for every_cell in column1:
    print(every_cell.value)

root = Tk()
root.title('Employee Record Management')
root.geometry('1300x650')
root.config(bg='#d4deec')

#FRAME FOR EMPLOYEE LABEL
frame = Frame(root, bg='#212530', borderwidth=2)
frame.grid(row=1, column=1, columnspan=15, ipadx=200, sticky=W)

#LEFT FRAME
frame1 = Frame(root, bg='#a9c3e6', borderwidth=3, relief="solid")
frame1.grid(row=2, column=1, columnspan=12, ipadx=10, sticky=W)

#BOtTOM FRAME IN VIEW DATA
frame2 = Frame(root, width=200)
frame2.grid(row=3, column=0, columnspan=50, sticky=W)

justtext = Text(root, width=35, height=13, bd=6, relief='sunken', font=("Courier", 15))
justtext.grid(row=2, column=10, ipady=31, padx=60)

def view_data():
        global tv1
        tv1 = ttk.Treeview(frame2, show='headings')
        
        treescrolly = Scrollbar(frame2, orient="vertical", command=tv1.yview)
        treescrollx = Scrollbar(frame2, orient="horizontal", command=tv1.xview)
        tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set)
        treescrollx.pack(side ="bottom",fill ="x")
        treescrolly.pack(side ="right",fill="y")  

        tv1['columns'] = ("Name", "Employee", "Email", "Department", "Designation", "Salary", "Present Days")
        tv1.column("Name", anchor=N, width=180)
        tv1.column("Employee",  anchor=N, width=180)
        tv1.column("Email", anchor=N, width=180)
        tv1.column("Department", anchor=N, width=180)
        tv1.column("Designation", anchor=N, width=180)
        tv1.column("Salary", anchor=N, width=180)
        tv1.column("Present Days", anchor=N, width=180)
        
        tv1.heading("Name", text="Name", anchor=N)
        tv1.heading("Employee", text="Employee", anchor=N)
        tv1.heading("Email", text="Email", anchor=N)
        tv1.heading("Department", text="Department", anchor=N)
        tv1.heading("Designation", text="Designation", anchor=N)
        tv1.heading("Salary", text="Salary", anchor=N)
        tv1.heading("Present Days", text="Present Days", anchor=N)
        

        for each_cell in range(2, (excel_activate.max_row)+1):
         tv1.insert(parent='', index="end", values=(excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value, excel_activate['E'+str(each_cell)].value, excel_activate['F'+str(each_cell)].value, excel_activate['G'+str(each_cell)].value, excel_activate['H'+str(each_cell)].value))
        tv1.pack()
        
# TO CREATE A FILE INSIDE EXCEL
student_records = []
     
def save():
    nm = nameentry.get()
    em = employeeentry.get()
    e = emailentry.get()
    dc = dept_var.get()
    mp = designation_var.get()
    se = salaryentry.get()
    pe = presententry.get()

    if nm == '' or em == '' or e == '' or dc == '' or mp == '' or se == '' or pe == '':
        messagebox.showerror('Error', 'Please fill in all fields')
    else:
        excel_con = load_workbook('finals.xlsx')
        excel_activate = excel_con.active
        row_num = 3
        while excel_activate.cell(row=row_num, column=1).value is not None:
            row_num += 1

        excel_con = load_workbook('finals.xlsx')
        excel_activate = excel_con.active
        row_num = 3
        while excel_activate.cell(row=row_num, column=1).value is not None:
            row_num +=1
            
        excel_activate.cell(row_num, column=1).value = nm
        excel_activate.cell(row_num, column=2).value = em
        excel_activate.cell(row_num, column=3).value = e
        excel_activate.cell(row_num, column=4).value = dc
        excel_activate.cell(row_num, column=5).value = mp
        excel_activate.cell(row_num, column=6).value = se
        excel_activate.cell(row_num, column=7).value = pe
        
        messagebox.showinfo('Database Record', 'Data Save Successfully!')
        excel_con.save('finals.xlsx')
        
        record = [nm, em, e, dc, se, pe]
        student_records.append(record)

        nameentry.delete(0, END)
        employeeentry.delete(0, END)
        emailentry.delete(0, END)
        deptcombo.set('')
        designationcombo.set('')
        salaryentry.delete(0, END)
        presententry.delete(0, END) 

def print_records(): #<- once the print btn click the data will print inside the Text frame
    justtext.delete('1.0', END)
    for record in student_records:
        justtext.insert(END, f'Name           : {record[0]}\n')
        justtext.insert(END, f'Employee ID    : {record[1]}\n')
        justtext.insert(END, f'Email          : {record[2]}\n')
        justtext.insert(END, f'Department     : {record[3]}\n')
        justtext.insert(END, f'Designation    : {record[4]}\n')
        justtext.insert(END, f'Salary         : {record[5]}\n')
        justtext.insert(END, f'Present Days   : {record[6]}\n')
        
        
        
#LABEL
welcome = Label(frame, text="EMPLOYEE REGISTRATION MANAGEMENT", font=('Courier',30), fg='#ffffff', bg='#212530')
welcome.grid(row=2, column=2, sticky=N, pady=20, padx=300)

#NAME
name = Label(frame1, text='Name: ', font=('Courier', 15), bg='#a9c3e6',)
name.grid(row=2, column=3, pady=20)

nameentry = Entry(frame1, font=('Courier', 13), bd=3)
nameentry.grid(row=2, column=4, ipady=3, padx=10)

#ID
employee = Label(frame1, text='Employee ID: ', font=('Courier', 13), bg='#a9c3e6',)
employee.grid(row=2, column=5)

employeeentry = Entry(frame1, font=('Courier', 13), bd=3)
employeeentry.grid(row=2, column=6, columnspan=5, pady=15, padx=10, ipady=3, sticky=W)

#EMAIL
email = Label(frame1, text='Email: ', font=('Courier', 15), bg='#a9c3e6')
email.grid(row=4, column=3)

emailentry = Entry(frame1, font=('Courier', 12), bd=3)
emailentry.grid(row=4, column=4, columnspan=5, padx=10, ipady=3, pady=3, sticky=W)

#DEPARTMENT
dept = Label(frame1, text='Department: ', font=('Courier', 14), bg='#a9c3e6',)
dept.grid(row=4, column=5)

dept_var = StringVar()
dept_list = ['HR', 'Engineering', 'Marketing', 'Planning', 'Admin']
deptcombo = ttk.Combobox(frame1, value=dept_list, textvariable=dept_var, font=('Courier', 10))
deptcombo.grid(row=4, column=6, ipadx=10, ipady=6, pady=10, padx=10, sticky=W)

#SALARY
salary = Label(frame1, text='Salary: ', font=('Courier', 15), bg='#a9c3e6',)
salary.grid(row=6, column=3)

salaryentry = Entry(frame1, font=('Courier', 10), bd=3)
salaryentry.grid(row=6, column=4, padx=10, ipady=3, sticky=W)


#DESIGNATION 
designation = Label(frame1, text='Designation: ', font=('Courier', 14), bg='#a9c3e6',)
designation.grid(row=6, column=5, padx=5)

designation_var = StringVar()
designation_list = ['IT', 'Data Analyst', 'Accounting Clerk', 'Electrical Enngineer', 'Planning Staff']
designationcombo = ttk.Combobox(frame1, value=designation_list, textvariable=designation_var, font=('Courier', 10))
designationcombo.grid(row=6, column=6, ipadx=10, ipady=3, pady=10, padx=10, sticky=W)

#PRESENT DAYS
present = Label(frame1, text='Present Days: ', font=('Courier', 14), bg='#a9c3e6',)
present.grid(row=8, column=3, padx=5, sticky=W)

presententry = Entry(frame1, font=('Courier', 10), bd=3)
presententry.grid(row=8, column=4, columnspan=1, ipadx=3, ipady=3, padx=10, pady=10, sticky='w')

#BUTTONS
#CREATE BTN
create = Button(frame1, text="Create", font=('Courier', 15), command=lambda:save())
create.grid(row=9, column=3, ipadx=4, pady=70)

#READ BTN
read = Button(frame1, text='Read', font=('Courier', 15), command=lambda:view_data())
read.grid(row=9, column=4, ipadx=7, pady=70, sticky=W)

#UPDATE BTN
update = Button(frame1, text='Delete', font=('Courier', 15))
update.grid(row=9, column=5, pady=70, columnspan=10)

#DELETE BTN
delete = Button(frame1, text="Update", font=('Courier', 15), command=lambda:refresh())
delete.grid(row=9, column=5, pady=70,)

#SHOW BTN
show = Button(frame1, text='Show', font=('Courier', 15), command=lambda:print_records())
show.grid(row=9, column=6, padx=50, pady=70, ipadx=7, sticky=E)


root.mainloop()