from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
from tkinter import messagebox, simpledialog
from tkinter import Tk, Button

from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook

import os 

excel_con = Workbook()
excel_con = load_workbook('finals.xlsx')
excel_activate = excel_con.active
column1= excel_activate['A']

for every_cell in column1:
    print(every_cell.value)
    
root = Tk()
root.title('Employee Record Management')
root.geometry('1300x630')
root.config(bg='#d4deec')

#FRAME FOR EMPLOYEE LABEL
employee_frame = Frame(root, bg='#212530', borderwidth=2)
employee_frame.grid(row=1, column=1, columnspan=15, ipadx=200, sticky=W)

#LEFT FRAME
left_frame = Frame(root, bg='#a9c3e6', borderwidth=3, relief="solid")
left_frame.grid(row=2, column=1, columnspan=12, ipadx=10, sticky=W)

#BOtTOM FRAME IN VIEW DATA
viewdata_frame= Frame(root, width=200)
viewdata_frame.grid(row=3, column=0, columnspan=50, sticky=W)

justtext = Text(root, width=35, height=13, bd=10, relief='sunken', font=("Courier", 15))
justtext.grid(row=2, column=10, ipady=31, padx=30)


student_records = []

def validate_employee_id(event):
    employee_id = employeeentry.get()
    if not employee_id.isdigit():
        messagebox.showerror('Error', 'Employee ID must be a number.')
        employeeentry.delete(0, END)

# TO CREATE A FILE INSIDE EXCEL     
def save():
    nm = nameentry.get()
    em = employeeentry.get()
    e = emailentry.get()
    dc = dept_var.get()
    mp = designationcombo.get() 
    se = salaryentry.get()  
    pe = presententry.get()

    if nm == '' or em == '' or e == '' or dc == '' or mp == '' or se == '' or pe == '':
        messagebox.showerror('Error', 'Please fill in all fields')
    else:
        excel_con = load_workbook('finals.xlsx')
        excel_activate = excel_con.active
        row_num = 3
        while excel_activate.cell(row=row_num, column=1).value is not None:
            row_num += 1

        excel_activate.cell(row_num, column=1).value = nm
        excel_activate.cell(row_num, column=2).value = em
        excel_activate.cell(row_num, column=3).value = e
        excel_activate.cell(row_num, column=4).value = dc
        excel_activate.cell(row_num, column=5).value = mp
        excel_activate.cell(row_num, column=6).value = se
        excel_activate.cell(row_num, column=7).value = pe

        messagebox.showinfo('Database Record', 'Data Saved Successfully!')
        excel_con.save('finals.xlsx')

        record = [nm, em, e, dc, mp, se, pe] 
        student_records.append(record)

        nameentry.delete(0, END)
        employeeentry.delete(0, END)
        emailentry.delete(0, END)
        deptcombo.set('')
        designationcombo.set('')
        salaryentry.delete(0, END)
        presententry.delete(0, END)

        tv1.insert(parent='', index="end", values=(nm, em, e, dc, mp, se, pe))

tv1 = None
read_button_clicked = False

def view_data():
    global tv1, read_button_clicked

    # Check if the read button has already been clicked
    if read_button_clicked:
        return

    # Set the flag variable to indicate the read button has been clicked
    read_button_clicked = True

    # Delete existing tv1 if it exists
    if tv1:
        tv1.pack_forget()
        tv1.destroy()

    tv1 = ttk.Treeview(viewdata_frame, show='headings', height=7)

    # Clear existing rows
    tv1.delete(*tv1.get_children())
    
    treescrolly = Scrollbar(viewdata_frame, orient="vertical", command=tv1.yview)
    treescrollx = Scrollbar(viewdata_frame, orient="horizontal", command=tv1.xview)
    tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set)
    treescrollx.pack(side="bottom", fill="x")
    treescrolly.pack(side="right", fill="y")

    tv1['columns'] = ("Name", "Employee", "Email", "Department", "Designation", "Salary", "Present Days")
    tv1.column("Name", anchor=N, width=180)
    tv1.column("Employee", anchor=N, width=180)
    tv1.column("Email", anchor=N, width=180)
    tv1.column("Department", anchor=N, width=180)
    tv1.column("Designation", anchor=N, width=180)
    tv1.column("Salary", anchor=N, width=180)
    tv1.column("Present Days", anchor=N, width=180)

    tv1.heading("Name", text="Name", anchor=N)
    tv1.heading("Employee", text="Employee", anchor=N)
    tv1.heading("Email", text="Email", anchor=N)
    tv1.heading("Department", text="Department", anchor=N)
    tv1.heading("Designation", text="Designation", anchor=N)
    tv1.heading("Salary", text="Salary", anchor=N)
    tv1.heading("Present Days", text="Present Days", anchor=N)

    for each_cell in range(2, (excel_activate.max_row)+1):
        tv1.insert(parent='', index="end", values=(excel_activate['A'+str(each_cell)].value, excel_activate['B'+str(each_cell)].value, excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value, excel_activate['E'+str(each_cell)].value, excel_activate['F'+str(each_cell)].value, excel_activate['G'+str(each_cell)].value, excel_activate['H'+str(each_cell)].value))
    
    tv1.pack(fill='both', expand=True)



#UPDATE FUNCTION
import openpyxl

def update_data():
    selected_item = tv1.focus()

    if selected_item == "":
        messagebox.showerror('Error', "No Item Selected")
        return

    # Update window is the new interface
    update_window = tk.Toplevel(root)
    update_window.title('Update Details')
    update_window.config(bg="#b3cde0")
    update_window.geometry('500x500')

    values = tv1.item(selected_item, "values")

    # TEXT AND WIDGETS
    name_label = tk.Label(update_window, text="Name:", font=('Courier', 15), bg="#b3cde0")
    name_label.grid(row=1, column=0, padx=30, pady=20, sticky=tk.W)
    name_entry = tk.Entry(update_window, font=('Courier', 12))
    name_entry.grid(row=1, column=1, pady=5, ipady=5)
    name_entry.insert(0, values[0])

    employee_id_label = tk.Label(update_window, text="Employee ID:", font=('Courier', 15), bg="#b3cde0")
    employee_id_label.grid(row=2, column=0, padx=20, pady=10, sticky=tk.W)
    employee_id_entry = tk.Entry(update_window, font=('Courier', 12))
    employee_id_entry.grid(row=2, column=1, padx=5, pady=5, ipady=5)
    employee_id_entry.insert(0, values[1])

    email_label = tk.Label(update_window, text="Email:", font=('Courier', 15), bg="#b3cde0")
    email_label.grid(row=3, column=0, padx=20, pady=10, sticky=tk.W)
    email_entry = tk.Entry(update_window, font=('Courier', 12))
    email_entry.grid(row=3, column=1, padx=5, pady=5, ipady=5)
    email_entry.insert(0, values[2])

    department_label = tk.Label(update_window, text="Department:", font=('Courier', 15), bg="#b3cde0")
    department_label.grid(row=4, column=0, padx=20, pady=10, sticky=tk.W)

    department_var = tk.StringVar()
    department_combobox = ttk.Combobox(update_window, textvariable=department_var, font=('Courier', 10))
    department_combobox['values'] = ['HR', 'Engineering', 'Marketing', 'Planning', 'Admin']
    department_combobox.grid(row=4, column=1, ipadx=10, padx=5, pady=5, ipady=5)
    department_combobox.set(values[3])

    designation_label = tk.Label(update_window, text="Designation:", font=('Courier', 15), bg="#b3cde0")
    designation_label.grid(row=5, column=0, padx=20, pady=10, sticky=tk.W)

    designation_var = tk.StringVar()
    designation_combobox = ttk.Combobox(update_window, textvariable=designation_var, font=('Courier', 10))
    designation_combobox['values'] = ['Chief Human Resources Officer', 
                    'HR Generalist', 
                    'HR Coordinator', 
                    'Mechanical Engineer', 
                    'Electrical Engineer', 
                    'Software Engineer', 
                    'Accounting Clerk', 
                    'Electrical Enngineer',
                    'Marketing Manager',
                    'Marketing Coordinator',
                    'Marketing Analyst',
                    'Planning Staff',
                    'Project Administrator',
                    'Office Manager']
    
    designation_combobox.grid(row=5, column=1, ipadx=10, padx=5, pady=5, ipady=5)
    designation_combobox.set(values[4])

    salary_label = tk.Label(update_window, text="Salary:", font=('Courier', 15), bg="#b3cde0")
    salary_label.grid(row=6, column=0, padx=20, pady=10, sticky=tk.W)
    salary_entry = tk.Entry(update_window, font=('Courier', 12))
    salary_entry.grid(row=6, column=1, padx=5, pady=5,ipady=5)
    salary_entry.insert(0, values[5])

    present_days_label = tk.Label(update_window, text="Present Days:", font=('Courier', 15), bg="#b3cde0")
    present_days_label.grid(row=7, column=0, padx=20, pady=10, sticky=tk.W)
    present_days_entry = tk.Entry(update_window, font=('Courier', 12))
    present_days_entry.grid(row=7, column=1, padx=5, pady=5, ipady=5)
    present_days_entry.insert(0, values[6])

    # SAVE BTN FOR THE UPDATE
    save_button = tk.Button(
        update_window,
        font=('Courier', 15),
        bg="#112236",
        fg="white",
        text="Save",
        command=lambda: save_updated_values(
            selected_item,
            name_entry.get(),
            employee_id_entry.get(),
            email_entry.get(),
            department_combobox.get(),
            designation_combobox.get(),
            salary_entry.get(),
            present_days_entry.get(),
            update_window
        )
    )
    save_button.grid(row=8, column=1, pady=50, sticky=tk.E)

    update_window.mainloop()

def save_updated_values(selected_item, name, employee_id, email, department, designation, salary, present_days, window):
    tv1.item(selected_item, values=(name, employee_id, email, department, designation, salary, present_days))
    messagebox.showinfo("Update", "Data updated successfully!")

    # Save the updated data to an Excel file
    workbook = openpyxl.load_workbook('finals.xlsx')
    worksheet = workbook.active

    # Get the row index of the selected item in the Excel file
    row_index = int(selected_item[1:]) + 1

    # Update the corresponding cells in the Excel file
    worksheet.cell(row=row_index, column=1).value = name
    worksheet.cell(row=row_index, column=2).value = employee_id
    worksheet.cell(row=row_index, column=3).value = email
    worksheet.cell(row=row_index, column=4).value = department
    worksheet.cell(row=row_index, column=5).value = designation
    worksheet.cell(row=row_index, column=6).value = salary
    worksheet.cell(row=row_index, column=7).value = present_days

    # Save the changes to the Excel file
    workbook.save('finals.xlsx')
    workbook.close()

    # Close the update window
    window.destroy()
    
#DELETE FUNCTION
def delete_data():
    selected_item = tv1.focus()
    if not selected_item:
        messagebox.showerror('Error', 'Please select a row to delete.')
        return
    
    # Confirm the deletion
    result = messagebox.askquestion('Delete', 'Are you sure you want to delete this record?')
    if result != 'yes':
        return

    index = int(tv1.index(selected_item))

    # Delete the row from the Treeview widget
    tv1.delete(selected_item)

    # Delete the row from the Excel file
    excel_activate.delete_rows(index + 2)

    # Save the changes back to the Excel file
    excel_con.save('finals.xlsx')
    messagebox.showinfo('Delete Data', 'Row deleted successfully.')

#To show the input in Text 
def print_records():
    justtext.delete('1.0', END)
    record = student_records[-1]  # Retrieve the last record from the list

    justtext.insert(END, f'Name           : {record[0]}\n')
    justtext.insert(END, f'Employee ID    : {record[1]}\n')
    justtext.insert(END, f'Email          : {record[2]}\n')
    justtext.insert(END, f'Department     : {record[3]}\n')
    justtext.insert(END, f'Designation    : {record[4]}\n')
    justtext.insert(END, f'Salary         : {record[5]}\n')
    justtext.insert(END, f'Present Days   : {record[6]}\n')

# Add this function to your code
def save_shortcut(event):
    save()

# Bind Ctrl+S to the save_shortcut function
root.bind('<Control-s>', save_shortcut)

def refresh_data(event=None):
    print_records()

# Bind Ctrl+R to refresh_data function
root.bind("<Control-r>", refresh_data)


#LABEL
welcome = Label(employee_frame, text="EMPLOYEE RECORD MANAGEMENT", font=('Courier',37), fg='#ffffff', bg='#212530')
welcome.grid(row=2, column=2, sticky=N, pady=20, padx=300)

#NAME
name = Label(left_frame, text='Name: ', font=('Courier', 15), bg='#a9c3e6',)
name.grid(row=2, column=3, pady=20)

nameentry = Entry(left_frame, font=('Courier', 13), bd=3)
nameentry.grid(row=2, column=4, ipady=3, padx=10)

#ID
employee = Label(left_frame, text='Employee ID: ', font=('Courier', 13), bg='#a9c3e6',)
employee.grid(row=2, column=5)

employeeentry = Entry(left_frame, font=('Courier', 13), bd=3)
employeeentry.grid(row=2, column=6, columnspan=5, pady=15, padx=10, ipady=3, sticky=W)
employeeentry.bind('<FocusOut>', validate_employee_id)

#EMAIL
email = Label(left_frame, text='Email: ', font=('Courier', 15), bg='#a9c3e6')
email.grid(row=4, column=3)

emailentry = Entry(left_frame, font=('Courier', 12), bd=3)
emailentry.grid(row=4, column=4, columnspan=5, padx=10, ipady=3, pady=3, sticky=W)

#DEPARTMENT
dept = Label(left_frame, text='Department: ', font=('Courier', 14), bg='#a9c3e6',)
dept.grid(row=4, column=5)

dept_var = StringVar()
dept_list = ['HR', 'Engineering', 'Marketing', 'Planning', 'Admin']
deptcombo = ttk.Combobox(left_frame, value=dept_list, textvariable=dept_var, font=('Courier', 10))
deptcombo.grid(row=4, column=6, ipadx=10, ipady=6, pady=10, padx=10, sticky=W)

#SALARY
salary = Label(left_frame, text='Salary: ', font=('Courier', 15), bg='#a9c3e6',)
salary.grid(row=6, column=3)

salaryentry = Entry(left_frame, font=('Courier', 10), bd=3)
salaryentry.grid(row=6, column=4, padx=10, ipady=3, sticky=W)


#DESIGNATION 
designation = Label(left_frame, text='Designation: ', font=('Courier', 14), bg='#a9c3e6',)
designation.grid(row=6, column=5, padx=5)

designation_var = StringVar()
designation_list = ['Chief Human Resources Officer', 
                    'HR Generalist', 
                    'HR Coordinator', 
                    'Mechanical Engineer', 
                    'Electrical Engineer', 
                    'Software Engineer', 
                    'Accounting Clerk', 
                    'Electrical Enngineer',
                    'Marketing Manager',
                    'Marketing Coordinator',
                    'Marketing Analyst',
                    'Planning Staff',
                    'Project Administrator',
                    'Office Manager']

designationcombo = ttk.Combobox(left_frame, value=designation_list, textvariable=designation_var, font=('Courier', 10))
designationcombo.grid(row=6, column=6, ipadx=10, ipady=3, pady=10, padx=10, sticky=W)

#PRESENT DAYS
present = Label(left_frame, text='Present Days: ', font=('Courier', 14), bg='#a9c3e6',)
present.grid(row=8, column=3, padx=5, sticky=W)

presententry = Entry(left_frame, font=('Courier', 10), bd=3)
presententry.grid(row=8, column=4, columnspan=1, ipadx=3, ipady=3, padx=10, pady=10, sticky='w')

#BUTTONS

style = ttk.Style()

# Button Clicked Style Configuration
style.map('PlainButton.TButton',
          background=[('active', 'royalblue')], 
          foreground=[('active', '#112236')]
          )

# CREATE BTN
create = ttk.Button(left_frame, text="Create", style='PlainButton.TButton', command=lambda: save())
create.grid(row=9, column=3, ipadx=8, ipady=5, pady=70)

# READ BTN
read = ttk.Button(left_frame, text='Read', style='PlainButton.TButton', command=lambda: view_data())
read.grid(row=9, column=4, ipadx=10, ipady=5, pady=70, sticky=W)

# DELETE BTN
delete = ttk.Button(left_frame, text='Delete', style='PlainButton.TButton', command=lambda: delete_data())
delete.grid(row=9, column=5, ipadx=10, ipady=5, pady=70, columnspan=10)

# UPDATE BTN
update = ttk.Button(left_frame, text="Update", style='PlainButton.TButton', command=lambda: update_data())
update.grid(row=9, column=5, ipadx=10, ipady=5, pady=70)

# SHOW BTN
show = ttk.Button(left_frame, text='Show', style='PlainButton.TButton', command=lambda: print_records())
show.grid(row=9, column=6, padx=40, pady=70, ipady=5, ipadx=7, sticky=E)

root.mainloop()